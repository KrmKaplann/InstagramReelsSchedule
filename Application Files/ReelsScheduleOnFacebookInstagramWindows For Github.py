import json
import os
import undetected_chromedriver as uc
import time
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import locale
import pyautogui
from datetime import datetime

locale.setlocale(locale.LC_TIME, "tr_TR.UTF-8")

def Driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    #options.add_argument("--incognito")
    #options.add_argument("--headless")  # Run the browser in headless (invisible) mode
    prefs = {"profile.default_content_setting_values.notifications": 2}

    options.add_experimental_option("prefs", prefs)
    options.add_argument("--start-maximized")  # Start the browser in full-screen mode
    driver = uc.Chrome(options=options)
    return driver

def Chrome_Tab(driver, link):
    driver.get(link)
    driver.maximize_window()
    driver.execute_script("document.body.style.zoom='75%'")

path = "PlaceHolder for SosyalMedyaHesap.xlsx path"  # Replace with the actual path
workbook = openpyxl.load_workbook(path)
DefaultPageSheet = workbook["Sayfa1"]

SelectedSocialMedia = "Facebook"  # Social media platform to process

#### Go to the row and column where the operation is located
StartColumn = 5
while True:
    Finder = DefaultPageSheet.cell(2, StartColumn).value
    if Finder == SelectedSocialMedia:
        break
    StartColumn += 1

### Add accounts to the list
TotalAccountListGmail = []
DefaultRow = 4

while True:
    Start = DefaultPageSheet.cell(DefaultRow, StartColumn).value
    if Start is None:
        break
    TotalAccountListGmail.append([Start, DefaultRow - 3])
    DefaultRow += 1

# Filter out items marked with '-'
TotalAccountListGmail = [item for item in TotalAccountListGmail if item[0] != '-']

TotalAccountList = []
DefaultRow = 4

for account_info in TotalAccountListGmail:
    index = account_info[1]  # Index value added to the first list
    Start = DefaultPageSheet.cell(index + 3, 2).value  # Calculate DefaultRow value from index
    TotalAccountList.append(Start)

print(TotalAccountList)

# Get the starting index from the user
starting_index = int(input("Enter the last completed index for Facebook main accounts (between 0 and {}): ".format(len(TotalAccountList)-1)))

# Perform the loop starting from the starting index
TotalAccountList = TotalAccountList[starting_index:]

InstagramAndFacebookAccountList = ["FacebookBusiness Name #1", "FacebookBusiness Name #2", "FacebookBusiness Name #3", "FacebookBusiness Name #3", "FacebookBusiness Name #4", "FacebookBusiness Name #5", "FacebookBusiness Name #6"]
print(InstagramAndFacebookAccountList)
# Get the starting index from the user
starting_index = int(input("Enter the last completed index for Facebook accounts (between 0 and {}): ".format(len(InstagramAndFacebookAccountList)-1)))

# Perform the loop starting from the starting index
InstagramAndFacebookAccountList = InstagramAndFacebookAccountList[starting_index:]

path = "PlaceHolder for FacebookAccounts.xlsx path"  # Replace with the actual path
workbook = openpyxl.load_workbook(path)
AllPagesWorksheet = workbook.sheetnames

for OneItem in TotalAccountListGmail:

    driver = Driver()
    link = "https://www.facebook.com"

    Chrome_Tab(driver, link)

    JsonFileName = DefaultPageSheet.cell(2, StartColumn).value + OneItem[0] + ".json"
    with open("PlaceHolder for JSON file path" + JsonFileName, "r") as file:  # Replace with the actual path
        cookies = json.load(file)

    for cookie in cookies:
        driver.add_cookie(cookie)

    time.sleep(3)

    driver.refresh()

    for index, OnePage in enumerate(TotalAccountList, start=1):
        print(f"[ {index} ] {OnePage}")

    StartRow = 5
    PostDict = {}

    for indexNO, OneAccountName in enumerate(AllPagesWorksheet[starting_index:]):
        StartRow = 5
        AllPostList = []
        while True:
            SpecialWorkSheet = workbook[OneAccountName]
            PostName = SpecialWorkSheet["B" + str(StartRow)].value
            StartRow += 1
            if PostName is None:
                break
            AllPostList.append(PostName)

        print(AllPostList)
        AllPostDefault = AllPostList
        # Get the starting index from the user
        starting_index = int(input(f"Enter the last completed index for the {OneAccountName} account (between 0 and {(len(AllPostList) - 1)}): "))

        # Perform the loop starting from the starting index
        AllPostList = AllPostList[starting_index:]
        PostStartNO = 0
        for StartRow, PostOne in enumerate(AllPostDefault, start=5):

            SpecialWorkSheet = workbook[OneAccountName]
            PostName = SpecialWorkSheet["B" + str(StartRow)].value
            Description = SpecialWorkSheet["C" + str(StartRow)].value  # Post description

            Date = SpecialWorkSheet["D" + str(StartRow)].value
            Date = datetime.strptime(Date, '%d.%m.%Y')  # Post date
            FormattedDate = Date.strftime("%d %B %A %Y")  # Formatted post date
            Day = Date.strftime("%d")  # Post day
            Month = Date.strftime("%B")  # Post month
            Time = str(SpecialWorkSheet["E" + str(StartRow)].value)  # Post time
            TimeSplitted = Time.split(":")
            Hour = str(TimeSplitted[0])  # Post hour
            Minutes = str(TimeSplitted[1])  # Post minutes
            ImmediateShare = SpecialWorkSheet["F" + str(StartRow)].value  # Immediate share flag

            # Create a dictionary for each post
            post_details = {
                "Description": Description,
                "Date": FormattedDate,
                "Day": Day,
                "Month": Month,
                "Time": Time,
                "Hour": Hour,
                "Minutes": Minutes,
                "ImmediateShare": ImmediateShare
            }

            # If OneAccountName is already in PostDict, get the existing dictionary and add the new post
            if OneAccountName in PostDict:
                PostDict[OneAccountName][PostName] = post_details
            else:
                # If OneAccountName is not in PostDict, create a new dictionary and add the new post
                PostDict[OneAccountName] = {PostName: post_details}

        for StartRow, PostOne in enumerate(AllPostList, start=5):

            wait = WebDriverWait(driver, 5)
            driver.implicitly_wait(15)  # seconds
            time.sleep(5)
            driver.get("https://business.facebook.com/latest/home?nav_ref=fb_web_pplus_settings_menu")  # <Facebook Business Suite URL>

            time.sleep(5)
            wait.until(EC.visibility_of_element_located((By.XPATH, "<Meta Business Suite element>")))  # <Meta Business Suite element>
            time.sleep(1)
            # Get all tabs/pages
            tabs = driver.window_handles

            # First tab/page (index 0)
            first_tab = tabs[0]
            driver.switch_to.window(first_tab)

            time.sleep(2)
            if indexNO == 0 and StartRow == 5:
                for i in range(3):
                    time.sleep(1)
                    pyautogui.hotkey('ctrl', '-')
                time.sleep(1)

            time.sleep(2)
            driver.find_element(By.XPATH, "<BizKitPresenceSelector element>").click()  # <BizKitPresenceSelector element>

            time.sleep(1)
            driver.find_element(By.XPATH, "<business entity search input>").send_keys(InstagramAndFacebookAccountList[indexNO])  # <business entity search input>
            time.sleep(3)
            driver.find_element(By.XPATH, "<business entity search result>").click()  # <business entity search result>
            time.sleep(2)

            if "#" in PostOne:
                driver.find_element(By.XPATH, "<Create Post button>").click()  # <Create Post button>

                while "#" in AllPostList[StartRow - 5 + PostStartNO]:
                    driver.find_element(By.XPATH, "<Add Photo button>").click()  # <Add Photo button>
                    driver.find_element(By.XPATH, "<Upload from Computer button>").click()  # <Upload from Computer button>
                    FileAddress = "PlaceHolder for Reels path" + OneAccountName + "/Reels/" + AllPostList[StartRow - 5 + PostStartNO] + ".jpeg"  # Replace with the actual path

                    time.sleep(3)
                    pyautogui.write(FileAddress)
                    time.sleep(2)
                    pyautogui.press('enter')

                    PostStartNO += 1

                driver.find_element(By.XPATH, "<combobox element>").click()  # <combobox element>

                StatusDict = {
                    "FacebookShare": True,
                    "InstagramShare": True,
                    "SavePreference": True  # Preference save flag
                }

                CheckBoxElements = driver.find_elements(By.XPATH, "<listbox elements>")  # <listbox elements>

                for index, CheckBoxElement in CheckBoxElements:
                    CheckBoxStatus = CheckBoxElement.get_attribute("aria-selected")
                    if CheckBoxStatus != StatusDict["FacebookShare"]:
                        CheckBoxElement.click()
                    if CheckBoxStatus != StatusDict["InstagramShare"]:
                        CheckBoxElement.click()
                    try:
                        if CheckBoxStatus != StatusDict["SavePreference"]:
                            CheckBoxElement.click()
                    except:
                        pass

                driver.find_element(By.XPATH, "<post description textbox>").send_keys("PostDict[OneAccountName][PostOne]['Description']")  # <post description textbox>

                driver.find_element(By.XPATH, "<schedule date and time input>").click()  # <schedule date and time input>

                for i in range(1, 3):
                    driver.find_element(By.XPATH, "<date input field>").send_keys(Keys.COMMAND + "a")  # <date input field>

                    # Convert the given date to a datetime object
                    date_obj = datetime.strptime(PostDict[OneAccountName][PostOne]["Date"], '%d %B %A %Y')  # Post date

                    # Convert to the new format
                    FormattedDate = date_obj.strftime('%d.%m.%Y')  # Formatted post date

                    driver.find_element(By.XPATH, "<date input field>").send_keys(FormattedDate)  # <date input field>

                    PostingTime = datetime.strptime(PostDict[OneAccountName][PostOne]["Time"], "%H:%M")  # Posting time

                    AmOrPm = PostingTime.strftime("%p")  # AM or PM
                    Hour = PostingTime.strftime("%I")  # Posting hour
                    Minutes = PostingTime.strftime("%M")  # Posting minutes

                    driver.find_element(By.XPATH, "<minute input field>").send_keys(Minutes)  # <minute input field>
                    driver.find_element(By.XPATH, "<hour input field>").send_keys(Hour)  # <hour input field>
                    driver.find_element(By.XPATH, "<AM/PM input field>").send_keys(AmOrPm)  # <AM/PM input field>

                    driver.find_element(By.XPATH, "<Schedule button>").click()  # <Schedule button>

            else:
                driver.find_element(By.XPATH, "<Create Reels Video button>").click()  # <Create Reels Video button>

                time.sleep(2)
                driver.find_element(By.XPATH, "<Add Video button>").click()  # <Add Video button>

                FileAddress = r"PlaceHolder for Reels video path" + "\u005C" + OneAccountName + "\u005C" + PostOne + ".mp4"  # Replace with the actual path
                # Convert all "/" characters to reverse ("\")

                print(FileAddress)

                time.sleep(2)
                pyautogui.write(FileAddress)
                time.sleep(2)
                pyautogui.press('enter')

                time.sleep(60)

                driver.find_element(By.XPATH, "<combobox element>").click()  # <combobox element>
                time.sleep(1)
                StatusDict = {
                    "FacebookShare": True,
                    "InstagramShare": True
                }

                CheckBoxElements = driver.find_elements(By.XPATH, "<listbox elements>")  # <listbox elements>

                for index, CheckBoxElement in enumerate(CheckBoxElements):
                    CheckBoxStatus = CheckBoxElement.get_attribute("aria-selected")
                    if CheckBoxStatus != StatusDict["FacebookShare"]:
                        CheckBoxElement.click()
                    if CheckBoxStatus != StatusDict["InstagramShare"]:
                        CheckBoxElement.click()

                time.sleep(2)
                driver.find_element(By.XPATH, "<post description textbox>").send_keys(PostDict[OneAccountName][PostOne]['Description'])  # <post description textbox>

                time.sleep(120)
                FileAddressCover = r"PlaceHolder for Cover image path" + "\u005C" + OneAccountName + "\u005C" + PostOne + "-" + "Kapak" ".jpg"  # Replace with the actual path

                # Check if the file exists
                if os.path.exists(FileAddressCover):
                    driver.find_element(By.XPATH, "<Upload Cover button>").click()  # <Upload Cover button>
                    time.sleep(4)
                    driver.find_element(By.XPATH, "<Upload Cover option>").click()  # <Upload Cover option>
                    time.sleep(3)
                    pyautogui.write(FileAddressCover)
                    time.sleep(2)
                    pyautogui.press('enter')
                    # Upload the file
                    time.sleep(5)

                else:
                    pass

                for i in range(2):
                    time.sleep(2)
                    for a in driver.find_elements(By.XPATH, "<Next buttons>"):  # <Next buttons>
                        time.sleep(3)
                        try:
                            a.click()
                        except:
                            pass
                time.sleep(2)

                time.sleep(20)
                driver.find_element(By.XPATH, "<Schedule button>").click()  # <Schedule button>
                time.sleep(1)
                driver.find_element(By.XPATH, "<date input field>").send_keys(Keys.COMMAND + "a")  # <date input field>
                time.sleep(0.5)
                # Convert the given date to a datetime object
                date_obj = datetime.strptime(PostDict[OneAccountName][PostOne]["Date"], '%d %B %A %Y')  # Post date

                # Convert to the new format
                FormattedDate = date_obj.strftime('%d.%m.%Y')  # Formatted post date

                driver.find_element(By.XPATH, "<date input field>").send_keys(Keys.CONTROL + "a")  # <date input field>
                time.sleep(1)
                driver.find_element(By.XPATH, "<date input field>").send_keys(FormattedDate)  # <date input field>

                PostingTime = datetime.strptime(PostDict[OneAccountName][PostOne]["Time"], "%H:%M")  # Posting time

                AmOrPm = PostingTime.strftime("%p")  # AM or PM
                HourPMandAM = PostingTime.strftime("%I")  # Posting hour in AM/PM format
                HourNormal = PostingTime.strftime("%H")  # Posting hour in 24-hour format
                Minutes = PostingTime.strftime("%M")  # Posting minutes

                driver.find_element(By.XPATH, "<minute input field>").send_keys(Minutes)  # <minute input field>
                driver.find_element(By.XPATH, "<hour input field>").send_keys(HourNormal)  # <hour input field>

                driver.find_element(By.XPATH, "<Schedule button>").click()  # <Schedule button>
